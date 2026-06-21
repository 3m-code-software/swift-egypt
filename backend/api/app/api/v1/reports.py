from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_operations_user, get_db
from app.models.customer import Customer
from app.models.shipment import Shipment, ShipmentStatus
from app.models.driver import Driver
from app.models.user import User
from app.schemas.report import CustomerReportItem, MonthlyReportItem, ReportResponse, StatusDistributionItem, DriverPerformanceItem

router = APIRouter(prefix="/reports", tags=["Reports"])


def _filter_by_date(query, start_date: date | None, end_date: date | None):
    if start_date:
        query = query.where(Shipment.created_at >= datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc))
    if end_date:
        query = query.where(Shipment.created_at <= datetime.combine(end_date, datetime.max.time(), tzinfo=timezone.utc))
    return query


@router.get("/summary", response_model=ReportResponse)
async def get_report_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_operations_user),
    start_date: date | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: date | None = Query(None, description="End date (YYYY-MM-DD)"),
):
    q = select(Shipment)
    q = _filter_by_date(q, start_date, end_date)
    result = await db.execute(q)
    all_shipments = result.scalars().all()

    months = {}
    for s in all_shipments:
        key = s.created_at.strftime("%Y-%m")
        if key not in months:
            months[key] = {"shipments": 0, "revenue": 0.0}
        months[key]["shipments"] += 1
        months[key]["revenue"] += s.final_price or s.estimated_price or 0

    monthly = sorted(
        [MonthlyReportItem(month=m, **d) for m, d in months.items()],
        key=lambda x: x.month,
    )

    status_map = {
        ShipmentStatus.delivered: ("Delivered", "#22c55e"),
        ShipmentStatus.in_transit: ("In Transit", "#6366f1"),
        ShipmentStatus.out_for_delivery: ("Out for Delivery", "#8b5cf6"),
        ShipmentStatus.pending: ("Pending", "#f59e0b"),
        ShipmentStatus.confirmed: ("Confirmed", "#3b82f6"),
        ShipmentStatus.picked_up: ("Picked Up", "#06b6d4"),
        ShipmentStatus.delayed: ("Delayed", "#ef4444"),
        ShipmentStatus.cancelled: ("Cancelled", "#6b7280"),
    }
    status_count = {}
    for s in all_shipments:
        status_count[s.status] = status_count.get(s.status, 0) + 1

    status_distribution = [
        StatusDistributionItem(name=status_map.get(st, (st.value, "#6b7280"))[0], value=count, color=status_map.get(st, ("Unknown", "#6b7280"))[1])
        for st, count in sorted(status_count.items(), key=lambda x: -x[1])
    ]

    driver_result = await db.execute(select(Driver).options(selectinload(Driver.user)))
    drivers = driver_result.scalars().all()
    driver_performance = []
    for d in drivers:
        if d.user:
            trips_result = await db.execute(select(func.count()).select_from(Shipment).where(Shipment.driver_id == d.id))
            trips = trips_result.scalar() or 0
            driver_performance.append(DriverPerformanceItem(name=d.user.full_name or "Unknown", rating=round(4.0 + (trips / 500) * 1.0, 1), trips=trips))

    customer_result = await db.execute(select(Customer).options(selectinload(Customer.user)))
    customers = customer_result.scalars().all()
    customer_report = []
    for c in customers:
        if c.user:
            cq = select(func.count(), func.coalesce(func.sum(Shipment.final_price), 0) + func.coalesce(func.sum(Shipment.estimated_price), 0)).select_from(Shipment).where(Shipment.customer_id == c.id)
            cq = _filter_by_date(cq, start_date, end_date)
            c_res = await db.execute(cq)
            total_shipments, total_revenue = c_res.one()
            total_revenue = float(total_revenue)
            customer_report.append(CustomerReportItem(
                customer_name=c.user.full_name or "Unknown",
                company_name=c.company_name,
                total_shipments=total_shipments or 0,
                total_revenue=round(total_revenue, 2),
                avg_shipment_value=round(total_revenue / total_shipments, 2) if total_shipments else 0,
            ))

    customer_report.sort(key=lambda x: -x.total_shipments)

    return ReportResponse(
        monthly=monthly,
        status_distribution=status_distribution,
        driver_performance=driver_performance,
        customers=customer_report,
    )


@router.get("/export/excel")
async def export_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_operations_user),
    start_date: date | None = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: date | None = Query(None, description="End date (YYYY-MM-DD)"),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    q = select(Shipment).order_by(Shipment.created_at.desc())
    q = _filter_by_date(q, start_date, end_date)
    result = await db.execute(q)
    shipments = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments Report"
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Swift Egypt - Shipments Report ({datetime.now().strftime('%Y-%m-%d')})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["Tracking #", "Status", "Service Type", "Sender", "Recipient", "Weight", "Price", "Created At"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, s in enumerate(shipments, 3):
        ws.cell(row=i, column=1, value=s.tracking_number)
        ws.cell(row=i, column=2, value=s.status.value if hasattr(s.status, 'value') else s.status)
        ws.cell(row=i, column=3, value=s.service_type.value if hasattr(s.service_type, 'value') else s.service_type)
        ws.cell(row=i, column=4, value=s.sender_name)
        ws.cell(row=i, column=5, value=s.recipient_name)
        ws.cell(row=i, column=6, value=s.weight or 0)
        price = s.final_price or s.estimated_price or 0
        ws.cell(row=i, column=7, value=float(price))
        ws.cell(row=i, column=8, value=s.created_at.strftime("%Y-%m-%d") if s.created_at else "")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=shipments_report_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )
