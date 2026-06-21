import io
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, Query, UploadFile
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_branch_manager_user, get_current_active_user, get_db, get_operations_user
from app.models.batch import Batch, BatchOrder, BatchStatus, OrderStatus
from app.models.branch import Branch
from app.models.driver import Driver
from app.models.seller import Seller
from app.models.user import User, UserRole
from app.schemas.batch import (BatchApproveRequest, BatchListResponse,
                               BatchOrderResponse, BatchOrderUpdate,
                               BatchRejectRequest, BatchResponse)

router = APIRouter(prefix="/batches", tags=["Batches"])


def generate_batch_number() -> str:
    now = datetime.now(timezone.utc)
    suffix = str(uuid.uuid4()).split("-")[0][:6].upper()
    return f"BATCH-{now.strftime('%Y%m%d')}-{suffix}"


@router.post("/upload")
async def upload_batch(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    seller_result = await db.execute(
        select(Seller).where(Seller.user_id == current_user.id)
    )
    seller = seller_result.scalar_one_or_none()
    if not seller and current_user.role not in [UserRole.admin, UserRole.operations]:
        from app.core.exceptions import BadRequestException
        raise BadRequestException("Seller profile not found")

    if current_user.role == UserRole.seller:
        seller_id = seller.id
    else:
        seller_id_str = None

    content = await file.read()
    workbook = None

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
    except Exception:
        from app.core.exceptions import BadRequestException
        raise BadRequestException("Invalid Excel file. Please upload a valid .xlsx file")

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if not rows:
        from app.core.exceptions import BadRequestException
        raise BadRequestException("Excel file is empty")

    expected_headers = ["customer_name", "customer_phone", "address", "product_name", "quantity", "product_price", "shipping_cost"]
    header_row = [str(cell).strip().lower() if cell else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]

    if not any(h in expected_headers for h in header_row):
        pass

    batch_number = generate_batch_number()
    batch = Batch(
        seller_id=seller.id if current_user.role == UserRole.seller else uuid.UUID(seller_id_str) if seller_id_str else seller.id,
        batch_number=batch_number,
        status=BatchStatus.pending,
        file_name=file.filename,
    )
    db.add(batch)
    await db.flush()

    total_amount = 0.0
    order_count = 0

    for row in rows:
        if not row or not row[0]:
            continue

        customer_name = str(row[0]) if row[0] else ""
        customer_phone = str(row[1]) if len(row) > 1 and row[1] else ""
        address = str(row[2]) if len(row) > 2 and row[2] else ""
        product_name = str(row[3]) if len(row) > 3 and row[3] else None
        quantity = int(row[4]) if len(row) > 4 and row[4] else 1
        product_price = float(row[5]) if len(row) > 5 and row[5] else 0.0
        shipping_cost = float(row[6]) if len(row) > 6 and row[6] else 0.0
        customer_phone2 = str(row[7]) if len(row) > 7 and row[7] else None
        province = str(row[8]) if len(row) > 8 and row[8] else None
        city = str(row[9]) if len(row) > 9 and row[9] else None
        notes = str(row[10]) if len(row) > 10 and row[10] else None

        total = (product_price + shipping_cost) * quantity

        order = BatchOrder(
            batch_id=batch.id,
            customer_name=customer_name,
            customer_phone=customer_phone,
            customer_phone2=customer_phone2,
            address=address,
            province=province,
            city=city,
            product_name=product_name,
            quantity=quantity,
            product_price=product_price,
            shipping_cost=shipping_cost,
            total=total,
            notes=notes,
        )
        db.add(order)
        total_amount += total
        order_count += 1

    batch.total_orders = order_count
    batch.total_amount = total_amount

    if seller:
        seller.total_orders += order_count

    await db.commit()

    return {
        "message": "Batch uploaded successfully",
        "batch_id": str(batch.id),
        "batch_number": batch.batch_number,
        "total_orders": order_count,
        "total_amount": total_amount,
    }


@router.get("/")
async def list_batches(
    status: str | None = Query(None),
    seller_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_operations_user),
):
    query = select(Batch).options(
        selectinload(Batch.seller).selectinload(Seller.user),
        selectinload(Batch.branch),
    ).order_by(Batch.created_at.desc())

    if status:
        query = query.where(Batch.status == status)
    if seller_id:
        query = query.where(Batch.seller_id == seller_id)

    result = await db.execute(query)
    batches = result.scalars().all()

    return [
        {
            "id": str(b.id),
            "seller_id": str(b.seller_id),
            "batch_number": b.batch_number,
            "status": b.status.value,
            "total_orders": b.total_orders,
            "total_amount": b.total_amount,
            "commission_percent": b.commission_percent,
            "seller_name": b.seller.user.full_name if b.seller and b.seller.user else None,
            "file_name": b.file_name,
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in batches
    ]


@router.get("/{batch_id}")
async def get_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_operations_user),
):
    result = await db.execute(
        select(Batch)
        .options(
            selectinload(Batch.seller).selectinload(Seller.user),
            selectinload(Batch.orders),
            selectinload(Batch.branch),
        )
        .where(Batch.id == batch_id)
    )
    b = result.scalar_one_or_none()
    if not b:
        from app.core.exceptions import NotFoundException
        raise NotFoundException("Batch not found")

    return {
        "id": str(b.id),
        "seller_id": str(b.seller_id),
        "branch_id": str(b.branch_id) if b.branch_id else None,
        "batch_number": b.batch_number,
        "status": b.status.value,
        "total_orders": b.total_orders,
        "total_amount": b.total_amount,
        "commission_percent": b.commission_percent,
        "commission_amount": b.commission_amount,
        "notes": b.notes,
        "reviewed_by": str(b.reviewed_by) if b.reviewed_by else None,
        "reviewed_at": b.reviewed_at.isoformat() if b.reviewed_at else None,
        "file_name": b.file_name,
        "seller_name": b.seller.user.full_name if b.seller and b.seller.user else None,
        "orders": [
            {
                "id": str(o.id),
                "batch_id": str(o.batch_id),
                "customer_name": o.customer_name,
                "customer_phone": o.customer_phone,
                "customer_phone2": o.customer_phone2,
                "address": o.address,
                "province": o.province,
                "city": o.city,
                "product_name": o.product_name,
                "quantity": o.quantity,
                "product_price": o.product_price,
                "shipping_cost": o.shipping_cost,
                "commission": o.commission,
                "total": o.total,
                "notes": o.notes,
                "status": o.status.value if o.status else "pending",
                "delivery_notes": o.delivery_notes,
                "returned_reason": o.returned_reason,
                "collected_amount": o.collected_amount,
                "latitude": o.latitude,
                "longitude": o.longitude,
                "assigned_agent_id": str(o.assigned_agent_id) if o.assigned_agent_id else None,
                "created_at": o.created_at.isoformat() if o.created_at else None,
                "updated_at": o.updated_at.isoformat() if o.updated_at else None,
            }
            for o in b.orders
        ],
        "created_at": b.created_at.isoformat() if b.created_at else None,
        "updated_at": b.updated_at.isoformat() if b.updated_at else None,
    }


@router.put("/{batch_id}/order/{order_id}")
async def update_order(
    batch_id: str,
    order_id: str,
    data: BatchOrderUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_branch_manager_user),
):
    result = await db.execute(
        select(BatchOrder).where(BatchOrder.id == order_id, BatchOrder.batch_id == batch_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        from app.core.exceptions import NotFoundException
        raise NotFoundException("Order not found")

    update_data = data.model_dump(exclude_none=True)
    for key, value in update_data.items():
        setattr(order, key, value)

    await db.flush()

    batch_result = await db.execute(select(Batch).where(Batch.id == batch_id))
    batch = batch_result.scalar_one()
    batch.updated_at = datetime.now(timezone.utc)

    total = sum(
        (o.product_price + o.shipping_cost) * o.quantity
        for o in (await db.execute(select(BatchOrder).where(BatchOrder.batch_id == batch_id))).scalars().all()
    )
    batch.total_amount = total
    await db.commit()

    return {"message": "Order updated successfully"}


@router.post("/{batch_id}/approve")
async def approve_batch(
    batch_id: str,
    data: BatchApproveRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_branch_manager_user),
):
    result = await db.execute(
        select(Batch).options(selectinload(Batch.orders)).where(Batch.id == batch_id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        from app.core.exceptions import NotFoundException
        raise NotFoundException("Batch not found")

    batch.status = BatchStatus.approved
    batch.commission_percent = data.commission_percent
    batch.commission_amount = batch.total_amount * (data.commission_percent / 100)
    batch.reviewed_by = current_user.id
    batch.reviewed_at = datetime.now(timezone.utc)
    batch.notes = data.notes

    for order in batch.orders:
        order.status = OrderStatus.approved
        order.commission = (order.product_price + order.shipping_cost) * (data.commission_percent / 100)

    seller_result = await db.execute(select(Seller).where(Seller.id == batch.seller_id))
    seller = seller_result.scalar_one()
    seller.wallet_balance += batch.total_amount - batch.commission_amount

    await db.commit()

    return {"message": "Batch approved successfully", "batch_id": batch_id}


@router.post("/{batch_id}/reject")
async def reject_batch(
    batch_id: str,
    data: BatchRejectRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_branch_manager_user),
):
    result = await db.execute(select(Batch).where(Batch.id == batch_id))
    batch = result.scalar_one_or_none()
    if not batch:
        from app.core.exceptions import NotFoundException
        raise NotFoundException("Batch not found")

    batch.status = BatchStatus.rejected
    batch.notes = data.reason
    batch.reviewed_by = current_user.id
    batch.reviewed_at = datetime.now(timezone.utc)

    await db.commit()

    return {"message": "Batch rejected", "batch_id": batch_id}


# Seller self-service endpoints
@router.get("/my/list")
async def list_my_batches(
    status: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    seller_result = await db.execute(
        select(Seller).where(Seller.user_id == current_user.id)
    )
    seller = seller_result.scalar_one_or_none()
    if not seller:
        from app.core.exceptions import NotFoundException
        raise NotFoundException("Seller profile not found")

    query = select(Batch).where(Batch.seller_id == seller.id).order_by(Batch.created_at.desc())
    if status:
        query = query.where(Batch.status == status)

    result = await db.execute(query)
    batches = result.scalars().all()

    return [
        {
            "id": str(b.id),
            "batch_number": b.batch_number,
            "status": b.status.value,
            "total_orders": b.total_orders,
            "total_amount": b.total_amount,
            "commission_percent": b.commission_percent,
            "file_name": b.file_name,
            "created_at": b.created_at.isoformat() if b.created_at else None,
        }
        for b in batches
    ]


@router.post("/seller/register")
async def register_seller(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    existing = await db.execute(
        select(Seller).where(Seller.user_id == current_user.id)
    )
    if existing.scalar_one_or_none():
        from app.core.exceptions import ConflictException
        raise ConflictException("Seller profile already exists")

    seller = Seller(
        user_id=current_user.id,
        company_name=current_user.full_name,
    )
    db.add(seller)
    current_user.role = UserRole.seller
    await db.commit()

    return {"message": "Seller profile created successfully", "seller_id": str(seller.id)}
